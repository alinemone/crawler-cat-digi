import requests
import xlsxwriter
import openpyxl

wb = openpyxl.load_workbook('Example2.xlsx')
sheet = wb.active

page_number = 1
sort_by = 4

category_list = [
    "category-men-blazers-suits",
    "category-men-homewear",
    "category-men-jeans",
    "category-men-knitwear",
    "category-men-outwear",
    "category-men-shirts",
    "category-men-shorts",
    "category-men-socks-tights",
    "category-men-hoodies",
    "category-men-sportswear",
    "category-men-swimwear",
    "category-men-tops",
    "category-men-tee-shirts-and-polos",
    "category-men-trousers-jumpsuits",
    "category-men-underwear",
    "category-men-ankle-boots",
    "category-men-boots",
    "category-men-espadrilles",
    "category-men-flip-flops",
    "category-men-formal-shoes",
    "category-men-sandals",
    "category-men-slippers",
    "category-men-sport-shoes-",
    "category-casual-shoes-for-men",
    "category-men-driving-shoes",
    "category-men-backpacks",
    "category-men-belts",
    "category-men-eyewear",
    "category-men-bags",
    "category-men-gifts-sets",
    "category-men-shawl",
    "category-men-gloves-mittens",
    "category-men-headwear",
    "category-men-jewelry",
    "category-men-luggage",
    "category-men-silver-jewelry",
    "category-men-sport-accessories",
    "category-men-ties",
    "category-men-wallets-bags",
    "category-men-watches",
    "category-men-glass-accessories",
    "category-men-keyring-and-keychain",
    "category-men-watch-accessories",
    "category-men-rosary",
    "category-men-cloth-face-mask",
    "category-women-blazers-and-suits",
    "category-women-dresses",
    "category-women-homewear",
    "category-women-islamicwear",
    "category-women-jeans",
    "category-women-knitwear",
    "category-women-maternity",
    "category-women-outwear",
    "category-women-shirts",
    "category-women-shorts-",
    "category-women-skirts",
    "category-women-socks-and-tights",
    "category-women-sportwear",
    "category-women-hoodies",
    "category-women-swimwear",
    "category-women-tee-shirts-and-polos",
    "category-women-tops",
    "category-women-trousers-and-jumpsuits",
    "category-women-legging",
    "category-women-underwear",
    "category-women-tunic",
    "category-women-ankle-boots",
    "category-women-boots",
    "category-women-espadrilles",
    "category-women-flat-shoes",
    "category-women-flip-flops",
    "category-women-heeled-shoes",
    "category-women-high-boots",
    "category-women-sandals",
    "category-women-bag-and-shoes-set",
    "category-women-slippers",
    "category-women-sport-shoes-",
    "category-casual-shoes-for-women",
    "category-women-shoe-care-accessories",
    "category-women-boat-shoes",
    "category-women-backpacks",
    "category-women-eyewear",
    "category-women-gifts-and-sets",
    "category-women-bags",
    "category-women-gloves-and-mittens",
    "category-women-belts",
    "category-women-gold-jewelry",
    "category-women-jewelry",
    "category-women-silver-jewelry",
    "category-women-hair-accessories",
    "category-women-headwear",
    "category-women-scarves",
    "category-women-sport-accessories",
    "category-women-ties",
    "category-women-wallets-and-cosmetic-bags",
    "category-women-watches",
    "category-women-glass-accessories",
    "category-women-jewelry-essential",
    "category-women-watch-accessories",
    "category-women-rosary",
    "category-women-scarves-accessories",
    "category-women-cloth-face-mask",
    "category-kids-clothes",
    "category-kids-footwear",
    "category-kids-accessories",
    "category-girls-clothes",
    "category-girls-footwear",
    "category-girls-accessories",
    "category-boys-clothes",
    "category-boys-shoes",
    "category-boys-accessories",
    "category-unikids-analouge-watches",
    "category-unikids-socks",
    "category-unikids-digital-watches",
    "category-unikids-homewear-sets",
    "category-unikids-homewear-tops-t-shirts",
    "category-unikids-homewear-bottom",
    "category-unikids-sport-trousers-and-jumpsuits",
    "category-girls-sport-t-shirts-polos1",
    "category-unikids-sport-shoes",
    "category-unikids-sport-sweatshirts-hoodies",
    "category-unikids-tracksuits-sets",
    "category-unikids-bracelets",
    "category-unikids-pendant",
    "category-unikids-trousers-jumpsuits",
    "category-unikids-gloves",
    "category-unikids-belts",
    "category-unikids-jeans",
    "category-unikids-tshirts-polos",
    "category-unikids-headwear",
    "category-unikids-shorts",
    "category-unikids-wallets",
    "category-unikids-sweatshirts",
    "category-unikids-backpacks",
    "category-unikids-bags",
    "category-unikids-glasses",
    "category-unikids-luggage",
    "category-kids-glass-accessories",
    "category-kids-face-masks",
    "category-uni-shoe-care-accessories",
    "category-uni-sport-shoes",
    "category-uni-backpacks",
    "category-uni-gifts-and-sets",
    "category-uni-socks",
    "category-uni-bags",
    "category-uni-headwear",
    "category-uni-eyewear",
    "category-uni-jewelry",
    "category-uni-jewelry-essential",
    "category-uni-luggage",
    "category-uni-shawl",
    "category-uni-silver-jewelry",
    "category-uni-sport-accessories",
    "category-uni-tech-accessories",
    "category-uni-wallets-bags",
    "category-uni-watches",
    "category-glass-accessories",
    "category-watch-accessories",
    "category-rosary",
    "category-uni-gold-jewelry",
    "category-uni-gloves",
    "category-unisex-ties",
    "category-cloth-face-mask",
    "category-badge"

]
workbook = xlsxwriter.Workbook('Example2.xlsx')
worksheet = workbook.add_worksheet()

row = 0 #sheet.max_row

print(row)

for category in category_list:
    for i in range(1, 30):
        response = requests.get(
            f'https://www.digikala.com/ajax/search/{category}/?has_selling_stock=1&/pageno={i}&sortby={sort_by}',
            cookies={
                "PHPSESSID": "62479tk3hg6q2sduhe142ktuhsagn8ohs5alt1b54ogg2hhsk9sbsug40d4dqvb6",
            })
        if response.status_code == 200:
            response_dict = dict(response.json())
            data = dict(response_dict.get('data'))
            click_impression = list(data.get('click_impression'))

            for item in click_impression:
                print(item['name'])
                worksheet.write_string(row, 0, str(item['supply_category'][1]))
                worksheet.write_string(row, 1, item['name'])

                row += 1

        else:
            break

workbook.close()
print("پایان")
